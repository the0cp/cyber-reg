import openpyxl

contents=[]

def read_txt():
    file_txt=open(r'../sheet.txt','r',encoding='utf-8')
    for i in file_txt:
        contents.append(i.split())
    file_txt.close()
    print(contents)

def write_excel():
    wb=openpyxl.Workbook()
    ws=wb.create_sheet(u'sheet')
    for i,content in enumerate(contents):
        for j in range(len(content)):
            ws.cell(i+1,j+1,content[j])
    wb.save('../sheet.xlsx')

read_txt()
write_excel()
